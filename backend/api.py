import tempfile
import os
from io import BytesIO
from typing import Any, Optional

from fastapi import FastAPI, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from utils import slugify

from services.excel_reader import load_workbook_model
from services.type_detector import detect_workbook_types
from services.pk_detector import detect_primary_keys, get_pk_candidates
from services.fk_detector import detect_foreign_keys
from services.database_connection import get_connection
from services.db_creator import create_tables
from services.db_crud import (
    get_table_schema, get_table_rows,
    insert_row, update_row, delete_row,
    get_row_references, delete_row_cascade,
)
from services.sessions import save_session, load_session
from models.relational.relational_db import RelationalDB
from models.relational.table import Table as RelTable
from models.relational.table_column import TableColumn
from models.relational.table_row import TableRow

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Modèles de requête pour /create ──────────────────────────────────────────

class ForeignKeyRef(BaseModel):
    refTable: str
    refColumn: str

class CreateColumn(BaseModel):
    name: str
    type: str
    isPrimaryKey: bool = False
    foreignKey: Optional[ForeignKeyRef] = None


class CreateTable(BaseModel):
    tableName: str
    columns: list[CreateColumn]
    rows: list[list[Any]]


class CreatePayload(BaseModel):
    tables: list[CreateTable]

@app.post("/parse")
async def parse_excel(file: UploadFile):

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    workbook = load_workbook_model(tmp_path)
    detect_workbook_types(workbook)
    detect_primary_keys(workbook)

    # Build relational model for FK detection
    relational_db = RelationalDB(name="db")
    for excel_table in workbook.get_all_tables():
        rel_table = RelTable(name=excel_table.name)
        for excel_col in excel_table.get_columns():
            rel_col = TableColumn(name=excel_col.name, relational_type=None)
            rel_col.is_primary_key = excel_col.is_primary_key
            rel_table.add_column(rel_col)
        for excel_row in excel_table.get_rows():
            row_dict = {
                excel_col.name: excel_row.cells[excel_col.index].value
                for excel_col in excel_table.get_columns()
            }
            rel_table.add_row(TableRow(row_dict))
        relational_db.add_table(rel_table)

    detect_foreign_keys(workbook, relational_db)

    # Build FK lookup: (table_name, col_name) → {refTable, refColumn}
    fk_map = {}
    for rel_table in relational_db.get_tables():
        for rel_col in rel_table.get_columns():
            if rel_col.foreign_key is not None:
                fk_map[(rel_table.name, rel_col.name)] = {
                    "refTable": rel_col.foreign_key.ref_table,
                    "refColumn": rel_col.foreign_key.ref_column,
                }

    os.unlink(tmp_path)

    sheets = []

    for worksheet in workbook.get_worksheets():

        tables = []

        for table in worksheet.get_tables():
            candidates = get_pk_candidates(table)

            columns = []
            for col in table.get_columns():
                pk_score = candidates.get(col.name, 0)
                columns.append({
                    "name": col.name,
                    "type": col.detected_type.value if col.detected_type else "STRING",
                    "isPrimaryKey": col.is_primary_key,
                    "isPkCandidate": col.name in candidates,
                    "pkScore": pk_score,
                    "foreignKey": fk_map.get((table.name, col.name)),
                })

            rows = [
                [cell.value for cell in row.get_cells()]
                for row in table.get_rows()
            ]

            tables.append({
                "name": table.name,
                "columns": columns,
                "rows": rows
            })

        sheets.append({
            "name": worksheet.name,
            "tables": tables
        })

    return {"sheets": sheets}


@app.post("/create")
def create_database(payload: CreatePayload):

    if not payload.tables:
        raise HTTPException(status_code=400, detail="Aucune table à créer.")

    try:
        conn = get_connection()
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=f"Connexion à la base de données impossible : {exc}"
        )

    try:
        tables = [table.model_dump() for table in payload.tables]
        results = create_tables(conn, tables)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la création des tables : {exc}"
        )
    finally:
        conn.close()

    return {"created": results}


# ─── CRUD generique ───────────────────────────────────────────────────────────

def _get_conn():
    try:
        return get_connection()
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Connexion impossible : {exc}")

def _coerce_pk(value: str, schema: list[dict]):
    pk = next((c for c in schema if c["isPrimaryKey"]), None)
    if pk and pk["type"] == "INT":
        try:
            return int(value)
        except ValueError:
            pass
    return value


@app.get("/tables/{name}/rows")
def read_rows(name: str):
    conn = _get_conn()
    try:
        columns = get_table_schema(conn, name)
        rows    = get_table_rows(conn, name)
        return {"columns": columns, "rows": rows}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()


@app.post("/tables/{name}/rows")
def create_row(name: str, data: dict[str, Any]):
    conn = _get_conn()
    try:
        row = insert_row(conn, name, data)
        return {"row": row}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()


@app.put("/tables/{name}/rows/{pk_value}")
def update_row_endpoint(name: str, pk_value: str, data: dict[str, Any]):
    conn = _get_conn()
    try:
        schema  = get_table_schema(conn, name)
        pk_col  = next((c["name"] for c in schema if c["isPrimaryKey"]), None)
        if not pk_col:
            raise HTTPException(status_code=400, detail="Aucune clé primaire trouvée.")
        coerced = _coerce_pk(pk_value, schema)
        row = update_row(conn, name, pk_col, coerced, data)
        if row is None:
            raise HTTPException(status_code=404, detail="Ligne introuvable.")
        return {"row": row}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()


@app.get("/tables/{name}/rows/{pk_value}")
def read_single_row(name: str, pk_value: str):
    """Retourne une seule ligne par sa valeur de PK."""
    conn = _get_conn()
    try:
        schema = get_table_schema(conn, name)
        pk_col = next((c["name"] for c in schema if c["isPrimaryKey"]), None)
        if not pk_col:
            raise HTTPException(status_code=400, detail="Aucune clé primaire trouvée.")
        coerced = _coerce_pk(pk_value, schema)
        from services.db_crud import _ident, _serialize_dict
        from utils import slugify
        cursor = conn.cursor()
        cursor.execute(
            f'SELECT * FROM {_ident(name)} WHERE {_ident(pk_col)} = %s LIMIT 1',
            (coerced,)
        )
        row_data = cursor.fetchone()
        cursor.close()
        if row_data is None:
            raise HTTPException(status_code=404, detail="Ligne introuvable.")
        col_names = [c["name"] for c in schema]
        row = _serialize_dict(dict(zip(col_names, row_data)))
        return {"row": row, "columns": schema}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()


@app.get("/tables/{name}/rows/{pk_value}/references")
def get_references(name: str, pk_value: str):
    """Retourne les lignes dépendantes de cette valeur de PK dans les tables FK."""
    conn = _get_conn()
    try:
        schema  = get_table_schema(conn, name)
        pk_col  = next((c["name"] for c in schema if c["isPrimaryKey"]), None)
        if not pk_col:
            raise HTTPException(status_code=400, detail="Aucune clé primaire trouvée.")
        coerced = _coerce_pk(pk_value, schema)
        refs    = get_row_references(conn, name, pk_col, coerced)
        return {"references": refs, "pkCol": pk_col}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()


@app.delete("/tables/{name}/rows/{pk_value}")
def delete_row_endpoint(name: str, pk_value: str, cascade: bool = False):
    conn = _get_conn()
    try:
        schema  = get_table_schema(conn, name)
        pk_col  = next((c["name"] for c in schema if c["isPrimaryKey"]), None)
        if not pk_col:
            raise HTTPException(status_code=400, detail="Aucune clé primaire trouvée.")
        coerced = _coerce_pk(pk_value, schema)
        deleted = delete_row_cascade(conn, name, pk_col, coerced) if cascade \
                  else delete_row(conn, name, pk_col, coerced)
        if not deleted:
            raise HTTPException(status_code=404, detail="Ligne introuvable.")
        return {"deleted": True}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()



# ─── Sessions ─────────────────────────────────────────────────────────────────

class SessionPayload(BaseModel):
    tables: list[str]
    config: dict = {}

@app.post("/sessions")
def create_session(payload: SessionPayload):
    """Sauvegarde une session et retourne un code court (EXC-XXXX)."""
    conn = _get_conn()
    try:
        code = save_session(conn, payload.tables, payload.config)
        return {"code": code}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()

@app.get("/sessions/{code}")
def get_session(code: str):
    """Charge une session par son code."""
    conn = _get_conn()
    try:
        session = load_session(conn, code)
        if session is None:
            raise HTTPException(status_code=404, detail="Session introuvable. Vérifiez le code.")
        return session
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()


@app.get("/export/excel")
def export_excel(tables: str):
    """Exporte toutes les tables demandées dans un seul fichier Excel (un onglet par table)."""
    table_names = [t.strip() for t in tables.split(',') if t.strip()]
    conn = _get_conn()

    wb = Workbook()
    wb.remove(wb.active)  # supprime la feuille vide par défaut

    try:
        for name in table_names:
            columns = get_table_schema(conn, name)
            rows    = get_table_rows(conn, name)
            ws = wb.create_sheet(title=slugify(name)[:31])
            col_names = [c["name"] for c in columns]
            ws.append(col_names)
            for row in rows:
                ws.append([row.get(c) for c in col_names])
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="export.xlsx"'},
    )


def _sql_value(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return str(v)
    return "'" + str(v).replace("'", "''") + "'"


@app.get("/export/sql")
def export_sql(tables: str):
    """Exporte toutes les tables demandées en SQL (CREATE TABLE + INSERT INTO)."""
    table_names = [t.strip() for t in tables.split(',') if t.strip()]
    conn = _get_conn()

    TYPE_MAP = {
        "INT": "INTEGER", "FLOAT": "DOUBLE PRECISION", "STRING": "TEXT",
        "DATE": "DATE",   "BOOL": "BOOLEAN",           "MIXED":  "TEXT",
    }
    parts: list[str] = []

    try:
        for name in table_names:
            columns   = get_table_schema(conn, name)
            rows      = get_table_rows(conn, name)
            safe      = slugify(name)
            col_names = [c["name"] for c in columns]

            # CREATE TABLE
            col_defs = []
            for c in columns:
                defn = f'  "{c["name"]}" {TYPE_MAP.get(c["type"], "TEXT")}'
                if c["isPrimaryKey"]:
                    defn += " PRIMARY KEY"
                col_defs.append(defn)
            parts.append(f'CREATE TABLE "{safe}" (\n' + ',\n'.join(col_defs) + '\n);')

            # INSERT INTO
            if rows:
                col_list = ', '.join(f'"{c}"' for c in col_names)
                value_rows = [
                    '(' + ', '.join(_sql_value(row.get(c)) for c in col_names) + ')'
                    for row in rows
                ]
                parts.append(f'INSERT INTO "{safe}" ({col_list}) VALUES\n' + ',\n'.join(value_rows) + ';')
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        conn.close()

    content = '\n\n'.join(parts).encode('utf-8')
    return StreamingResponse(
        BytesIO(content),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="export.sql"'},
    )
