import re
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from models.excel.workbook import Workbook
from models.excel.worksheet import Worksheet
from models.excel.table import Table
from models.excel.row import Row
from models.excel.column import Column
from models.excel.cell import Cell
from errors import ExcelReadError
from constants import DATE_FORMATS


# ─── Utilitaires ──────────────────────────────────────────────────────────────

def clean_header(value):
    if value is None:
        return ""
    return str(value).lower().strip().replace(" ", "_")


HEADER_SCORE_THRESHOLD = 5


def looks_like_number(value):
    try:
        float(str(value).replace(",", "."))
        return True
    except ValueError:
        return False


def looks_like_date(value):
    for fmt in DATE_FORMATS:
        try:
            datetime.strptime(str(value).strip(), fmt)
            return True
        except ValueError:
            pass
    return False


TIME_PATTERN = re.compile(r'^\d{1,2}:\d{2}(:\d{2})?$')


def looks_like_time(value):
    return bool(TIME_PATTERN.match(str(value).strip()))


def looks_like_alnum_code(value):
    v = str(value).strip()
    return any(c.isdigit() for c in v) and any(c.isalpha() for c in v)


def get_value_type(value):
    from datetime import date, datetime as dt
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, (date, dt)):
        return "date"
    if isinstance(value, str):
        v = value.strip()
        if looks_like_date(v):
            return "date"
        if looks_like_time(v):
            return "time"
        if looks_like_number(v):
            return "numeric_str"
        if looks_like_alnum_code(v):
            return "alnum_code"
        return "str"
    return "unknown"


def dominant_type(values):
    types = [get_value_type(v) for v in values if v is not None]
    if not types:
        return "null"
    return max(set(types), key=types.count)


def is_sequential(values):
    cleaned = [v for v in values if v is not None]
    if len(cleaned) < 2:
        return False

    # Séquence numérique : 1,2,3...
    if all(isinstance(v, (int, float)) for v in cleaned):
        nums = sorted(cleaned)
        return all(nums[i+1] - nums[i] == nums[1] - nums[0] for i in range(len(nums)-1))

    # Séquence alphabétique : A,B,C...
    if all(isinstance(v, str) and len(v.strip()) == 1 for v in cleaned):
        letters = sorted(v.strip().upper() for v in cleaned)
        return all(ord(letters[i+1]) - ord(letters[i]) == 1 for i in range(len(letters)-1))

    # Pattern col1,col2,col3 ou colonne_1,colonne_2...
    if all(isinstance(v, str) for v in cleaned):
        import re
        pattern = re.compile(r'^(.+?)(\d+)$')
        matches = [pattern.match(str(v).strip().lower()) for v in cleaned]
        if all(m for m in matches):
            prefixes = {m.group(1) for m in matches}
            numbers = sorted(int(m.group(2)) for m in matches)
            if len(prefixes) == 1:
                return all(numbers[i+1] - numbers[i] == 1 for i in range(len(numbers)-1))

    return False


def is_row_empty(ws, row_index, max_col):
    return all(
        ws.cell(row=row_index, column=col).value is None
        or str(ws.cell(row=row_index, column=col).value).strip() == ""
        for col in range(1, max_col + 1)
    )


# ─── Détection de header ──────────────────────────────────────────────────────

def detect_has_header(first_row_values, data_rows_values):
    non_empty_first = [v for v in first_row_values if v is not None and str(v).strip() != ""]

    if not non_empty_first:
        return False

    # Sans données suivantes on ne peut pas trancher
    if not data_rows_values:
        return len(set(get_value_type(v) for v in non_empty_first)) == 1

    score = 0

    # ── Critère 1 : contraste de type par colonne (+3) ──────────────────────
    # Pour chaque colonne, on compare le type de la cellule ligne 1
    # au type dominant des données de cette colonne
    col_count = len(first_row_values)
    contrast_count = 0

    for col_i in range(col_count):
        first_val = first_row_values[col_i]
        col_data = [
            row[col_i] for row in data_rows_values
            if col_i < len(row) and row[col_i] is not None
        ]
        if not col_data:
            continue
        if get_value_type(first_val) != dominant_type(col_data):
            contrast_count += 1

    if col_count > 0 and contrast_count / col_count >= 0.5:
        score += 3

    # ── Critère 2 : ligne 1 homogène + données hétérogènes (+2) ─────────────
    first_types = {get_value_type(v) for v in non_empty_first} - {"null"}
    all_data_values = [
        v for row in data_rows_values
        for v in row
        if v is not None and str(v).strip() != ""
    ]
    data_types = {get_value_type(v) for v in all_data_values} - {"null"}

    if len(first_types) == 1 and len(data_types) > 1:
        score += 2

    # ── Critère 3 : valeurs séquentielles en ligne 1 (+2) ───────────────────
    if is_sequential(non_empty_first):
        score += 2

    # ── Critère 4 : toutes les valeurs de ligne 1 sont uniques (+1) ─────────
    normalized = [str(v).strip().lower() for v in non_empty_first]
    if len(normalized) == len(set(normalized)):
        score += 1

    # ── Critère 5 : aucun null en ligne 1 (+1) ──────────────────────────────
    if None not in first_row_values and all(str(v).strip() != "" for v in first_row_values):
        score += 1

    # ── Critère 6 : strings courtes en ligne 1 vs données (+1) ─────────────
    first_strings = [str(v) for v in non_empty_first if isinstance(v, str)]
    data_strings = [str(v) for v in all_data_values if isinstance(v, str)]
    if first_strings and data_strings:
        avg_first = sum(len(s) for s in first_strings) / len(first_strings)
        avg_data = sum(len(s) for s in data_strings) / len(data_strings)
        if avg_first < avg_data:
            score += 1

    # ── Critère 7 : valeurs ligne 1 absentes des données (+1) ───────────────
    data_set = {str(v).strip().lower() for v in all_data_values}
    if not any(str(v).strip().lower() in data_set for v in non_empty_first):
        score += 1

    # ── Critère 8 : format identifiant (+2) ──────────────────────────────────
    # Les noms de colonnes suivent typiquement snake_case ou ALL_CAPS
    # ex: nom, ville, employee_id, TOTAL → oui
    # ex: Farah, Omar, Paris → non (Title Case = donnée)
    import re
    identifier_pattern = re.compile(r'^([a-z][a-z0-9_]*|[A-Z][A-Z0-9_]*)$')
    str_values = [v for v in non_empty_first if isinstance(v, str)]
    if str_values and all(identifier_pattern.match(str(v).strip()) for v in str_values):
        score += 2

    return score >= HEADER_SCORE_THRESHOLD


# ─── Détection de blocs ───────────────────────────────────────────────────────

def detect_blocks(ws):
    blocks = []
    block_start = None

    for row_index in range(1, ws.max_row + 1):
        if not is_row_empty(ws, row_index, ws.max_column):
            if block_start is None:
                block_start = row_index
        else:
            if block_start is not None:
                blocks.append((block_start, row_index - 1))
                block_start = None

    if block_start is not None:
        blocks.append((block_start, ws.max_row))

    return blocks


# ─── Chargement d'un bloc ─────────────────────────────────────────────────────

def load_block(ws, table_name, start_row, end_row, ws_formulas=None):
    table = Table(name=table_name, index=0)

    first_row_values = [
        ws.cell(row=start_row, column=col).value
        for col in range(1, ws.max_column + 1)
    ]

    data_rows_values = [
        [ws.cell(row=r, column=col).value for col in range(1, ws.max_column + 1)]
        for r in range(start_row + 1, min(start_row + 4, end_row + 1))
    ]

    table.set_has_header(detect_has_header(first_row_values, data_rows_values))

    used_columns = [
        col_index for col_index in range(ws.max_column)
        if any(
            ws.cell(row=r, column=col_index + 1).value is not None
            for r in range(start_row, end_row + 1)
        )
    ]

    for local_index, col_index in enumerate(used_columns):
        if table.get_has_header():
            name = clean_header(first_row_values[col_index]) or f"column_{local_index + 1}"
        else:
            name = f"column_{local_index + 1}"

        column = Column(
            name=name,
            index=local_index,
            letter=ws.cell(row=start_row, column=col_index + 1).column_letter
        )
        table.add_column(column)

    data_start = start_row + 1 if table.get_has_header() else start_row

    for row_index in range(data_start, end_row + 1):
        row = Row(row_index)

        for local_index, col_index in enumerate(used_columns):
            value = ws.cell(row=row_index, column=col_index + 1).value

            formula = None
            if ws_formulas is not None:
                raw = ws_formulas.cell(row=row_index, column=col_index + 1).value
                if isinstance(raw, str) and raw.startswith("="):
                    formula = raw

            row.add_cell(Cell(
                value=value,
                row_index=row_index,
                column_index=local_index + 1,
                formula=formula
            ))

        table.add_row(row)

    return table


# ─── Chargement du classeur ───────────────────────────────────────────────────

def load_workbook_model(path):
    if path is None or str(path).strip() == "":
        raise ExcelReadError("Aucun fichier fourni")

    path = Path(path)

    if path.suffix.lower() != ".xlsx":
        raise ExcelReadError("Format invalide")

    if not path.exists():
        raise ExcelReadError(f"Fichier introuvable : {path}")

    try:
        wb = load_workbook(path, data_only=True)
        wb_formulas = load_workbook(path)
    except InvalidFileException:
        raise ExcelReadError(f"Le fichier '{path.name}' est corrompu ou n'est pas un fichier Excel valide")

    workbook = Workbook(file_name=path.name, file_path=str(path))

    for sheet_index, ws in enumerate(wb.worksheets):
        ws_formulas = wb_formulas.worksheets[sheet_index]
        worksheet = Worksheet(name=ws.title, index=sheet_index)
        pending_title = None

        for block_index, (start_row, end_row) in enumerate(detect_blocks(ws)):
            table_name = ws.title if block_index == 0 else f"{ws.title}_{block_index + 1}"
            table = load_block(ws, table_name, start_row, end_row, ws_formulas)

            if table.is_empty():
                cell_value = ws.cell(row=start_row, column=1).value
                if cell_value and str(cell_value).strip():
                    pending_title = str(cell_value).strip().lower().replace(" ", "_")
                continue

            if pending_title is not None:
                table.name = pending_title
                pending_title = None

            worksheet.add_table(table)

        workbook.add_worksheet(worksheet)

    return workbook
