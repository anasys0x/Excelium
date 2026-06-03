from openpyxl import load_workbook

from src.excelium.services.sql_mapper import map_all_columns_to_sql, create_table_sql


def open_workbook(path: str):
    wb = load_workbook(path)
    return wb

def get_worksheets(workbook):
    worksheet_names  = []
    ws = workbook.worksheets
    for sheet_name in ws:
        worksheet_names.append(sheet_name.title)
    return worksheet_names

def get_active_worksheet(workbook):
    return workbook.active

def get_headers(worksheet):
    cols = [cell.value for cell in worksheet[1]]
    return cols

def clean_headers(headers):
    header = [str(e).lower().strip().replace(" ", "_") for e in headers]
    return header

def get_data(worksheet):
    data = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        #On arrete des qu'une ligne entiere ne contient que des Nones (cases vides)
        if all(cell is None for cell in row):
            break
        data.append(list(row))
    return data

def get_column_values(data, column_index):
    values = []
    for row in data:
        values.append(row[column_index])
    return values




